#!/usr/bin/env python3
"""Generate detailed SciAi-Nova OS vs Microsoft Discovery comparison workbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent.parent / "docs" / "SciAi-Nova_vs_Microsoft_Discovery_Comparison.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
SUBHEADER_FILL = PatternFill("solid", fgColor="D6EAF8")
ACCENT_FILL = PatternFill("solid", fgColor="E8F6F3")
THIN = Side(style="thin", color="BDC3C7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_subheader_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, size=10)
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def write_table(ws, headers: list[str], rows: list[list], start_row: int = 1) -> int:
    cols = len(headers)
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header_row(ws, start_row, cols)
    r = start_row + 1
    for row_data in rows:
        for i, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
        r += 1
    return r


def set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_executive_summary(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    ws["A1"] = "SciAi-Nova OS vs Microsoft Discovery — Detailed Comparison"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A2"] = "Generated for TCS pharma R&D positioning | Sources: SciAi-Nova OS codebase + Microsoft Learn (Discovery)"
    ws.merge_cells("A2:E2")

    headers = ["Dimension", "SciAi-Nova OS", "Microsoft Discovery (Cloud)", "Microsoft Discovery App", "Notes"]
    rows = [
        [
            "Product type",
            "Self-hosted Scientific Data Fabric + Agent Operating System",
            "Enterprise Azure agentic R&D platform (GA)",
            "Free local desktop on-ramp (preview)",
            "App shares APIs/concepts with cloud platform",
        ],
        [
            "Primary deployment",
            "Docker Compose on Azure VM, AWS EC2, or local Mac",
            "Azure subscription — Discovery Studio + ARM resources",
            "Windows desktop — local compute",
            "",
        ],
        [
            "Target buyer",
            "TCS demos, pilots, partner PoCs, multi-cloud flexibility",
            "Enterprise R&D orgs committed to Azure at scale",
            "Individual researchers, students, quick experiments",
            "",
        ],
        [
            "Time to first demo",
            "Hours (git clone + docker compose + API keys)",
            "Weeks (Azure setup, workspace, Bookshelf, governance)",
            "Minutes (download + GitHub Copilot account)",
            "",
        ],
        [
            "Licensing / cost model",
            "No platform license; VM + OpenAI/Bedrock/Mistral API costs",
            "Consumption (billable user messages) + Azure infra + models",
            "Free download; consumes GitHub Copilot credits",
            "See Cost sheet",
        ],
        [
            "Agentic discovery loop",
            "Workflow orchestrator + ~24 real agent pipelines",
            "Discovery Engine — hypothesis tree, validation, long-term memory",
            "Same core agent concepts, local scale",
            "Microsoft deeper on scientific-method orchestration",
        ],
        [
            "Knowledge graph",
            "PostgreSQL + Neo4j; pattern + LLM extraction from uploads",
            "Bookshelf → Knowledge Bases via GraphRAG (vector + graph)",
            "Bookshelf-style indexing at local scale",
            "",
        ],
        [
            "HPC / simulation",
            "Not included (single VM)",
            "Supercomputer on AKS — simulations, tool workloads",
            "Local compute only",
            "Major Microsoft differentiator",
        ],
        [
            "Governance / compliance",
            "Audit, approvals, GxP UI — pilot grade",
            "Enterprise governance, certifications, SLAs",
            "Fixed safeguards; no enterprise certifications",
            "",
        ],
        [
            "Best pharma demo today",
            "MASH PDFs → fabric search → KG → Target Discovery workflow",
            "Literature synthesis, candidate ID, experiment design at enterprise scale",
            "Solo literature / agent exploration",
            "",
        ],
        [
            "Overall positioning",
            "Rapid deployable Agent OS you own and customize",
            "Enterprise thinking partner for full R&D program arc",
            "Frictionless individual entry to Discovery concepts",
            "Complementary, not direct substitute",
        ],
    ]
    write_table(ws, headers, rows, start_row=4)
    set_col_widths(ws, [22, 38, 38, 32, 28])


def sheet_architecture(wb: Workbook) -> None:
    ws = wb.create_sheet("Architecture")
    headers = ["Component", "SciAi-Nova OS", "Microsoft Discovery", "Parity", "Gap / Advantage"]
    rows = [
        ["Frontend", "React + Vite + Tailwind; 3D KG (ForceGraph3D)", "Discovery Studio (cloud UX); Discovery app (desktop)", "Partial", "Different UX; SciNova fully customizable"],
        ["API layer", "FastAPI REST + JWT", "Discovery REST APIs + Azure ARM", "Partial", "Both API-driven"],
        ["Primary database", "PostgreSQL 16", "Azure-managed storage + project resources", "Partial", ""],
        ["Vector store", "ChromaDB", "GraphRAG vector index (Bookshelf)", "Partial", "Microsoft GraphRAG more integrated"],
        ["Graph database", "Neo4j 5 (+ SQL graph source of truth)", "Knowledge graph in Bookshelf / long-term memory", "Partial", "Microsoft org-wide memory; SciNova project-scoped"],
        ["Job queue", "Redis + Celery", "Azure-native orchestration + Supercomputer jobs", "Partial", ""],
        ["File storage", "Local/docker volumes (uploads)", "Azure storage", "Partial", ""],
        ["LLM routing", "OpenAI, Bedrock Claude, Mistral SLM (configurable)", "Azure OpenAI / AI Foundry (e.g. GPT-class)", "Partial", "SciNova multi-cloud; MS Azure-centric"],
        ["Embeddings", "OpenAI text-embedding-3-small or Bedrock Titan", "Foundry / platform-managed", "Partial", "SciNova supports Bedrock-only AWS path"],
        ["Ingestion worker", "Celery worker container", "Platform-managed ingestion pipelines", "Partial", ""],
        ["Multi-tenancy", "Users + projects + workspace scoping", "Enterprise workspaces, RBAC, investigations", "Partial", "Microsoft stronger at enterprise scale"],
        ["Cloud portability", "Azure VM, AWS EC2, local Docker", "Azure only (cloud platform)", "SciNova advantage", "SciNova runs same stack on AWS Bedrock demo"],
        ["Source code access", "Full GitHub repo — fork and extend", "Closed platform; extend via tools/APIs", "SciNova advantage", ""],
        ["Desktop app", "Browser only (port 5173)", "Discovery app (Windows, preview)", "MS app only", ""],
        ["Container count (typical)", "7 (frontend, backend, celery, postgres, neo4j, redis, chroma)", "Many Azure services (AKS, storage, Foundry, etc.)", "N/A", "SciNova simpler ops on one VM"],
    ]
    write_table(ws, headers, rows)
    set_col_widths(ws, [24, 36, 36, 12, 32])


def sheet_knowledge_data(wb: Workbook) -> None:
    ws = wb.create_sheet("Knowledge & Data")
    headers = [
        "Capability", "SciAi-Nova OS", "Microsoft Discovery",
        "SciNova maturity", "Microsoft maturity", "Demo relevance",
    ]
    rows = [
        ["Document upload (PDF)", "Data Fabric — Celery pipeline", "Bookshelf ingestion", "Production (demo)", "GA", "High"],
        ["DOCX / TXT", "Parser-dependent; PDF primary path", "Supported in Bookshelf", "Partial", "GA", "Medium"],
        ["XLSX / CSV (assay data)", "Supported with QC stage", "Supported", "Good", "GA", "Medium"],
        ["Chunking strategy", "~512 tokens, overlap 64", "Platform-managed", "Good", "GA", "Medium"],
        ["Embeddings index", "ChromaDB per document", "GraphRAG dual index", "Good", "GA", "High"],
        ["Entity extraction", "Regex patterns + frontier LLM hybrid", "Platform + GraphRAG indexing", "Good", "GA", "High"],
        ["Relationship extraction", "Co-occurrence + LLM; fixed in recent release", "GraphRAG + agent reasoning", "Improving", "GA", "Medium"],
        ["Ontology mapping", "Light ontology_id on entities", "Curated KBs, standards alignment", "Basic", "Strong", "Low for demo"],
        ["Knowledge graph UI", "3D explorer; SQL/Neo4j/auto sources", "Studio + investigations", "Good", "GA", "High"],
        ["Project scoping", "X-Project-Id header; per-project docs", "Discovery projects / investigations", "Good", "GA", "High"],
        ["Semantic search", "Fabric search + RAG", "NL over Bookshelf + Engine", "Good", "GA", "High"],
        ["PubMed integration", "Yes (agent/evidence layer)", "External literature in discovery loop", "Good", "GA", "High"],
        ["KEGG / pathway APIs", "Yes", "Via tools / external data", "Good", "Likely", "Medium"],
        ["Federation of enterprise DBs", "Not native (upload/export)", "Virtual graphs / Fabric integration", "Gap", "Strong", "Low for demo"],
        ["Long-term org memory", "Per-deployment Postgres", "Discovery Engine long-term KG", "Gap", "Strong", "Medium"],
        ["GraphRAG", "No (separate vector + property graph)", "Yes — Microsoft Research technique", "Gap", "Strong", "Medium"],
        ["Evidence / citations", "Chunk excerpts + provenance JSON", "Explainable Engine outcomes", "Good", "Strong", "High"],
        ["Re-ingest / backfill", "Manual backfill script for relationships", "Platform-managed KB refresh", "Manual", "GA", "Medium"],
    ]
    write_table(ws, headers, rows)
    set_col_widths(ws, [28, 34, 34, 14, 16, 14])


def sheet_agents(wb: Workbook) -> None:
    ws = wb.create_sheet("Agents & Orchestration")
    headers = ["Capability", "SciAi-Nova OS", "Microsoft Discovery", "SciNova status", "Notes"]
    rows = [
        ["Agent marketplace", "~78 agents defined in seed data", "Domain agents across science/engineering", "~24 real pipelines", "MS agents are first-class product"],
        ["Discovery Engine equivalent", "Workflow orchestrator + agent executor", "Discovery Engine — cognitive orchestrator", "Functional", "MS formal hypothesis tree"],
        ["Multi-step workflows", "Target Discovery, Literature→Hypothesis, etc.", "Investigations + repeatable pipelines", "Good", ""],
        ["Live orchestration UI", "Animated agent strip, progress polling", "Discovery Studio", "Good", ""],
        ["Human approval gates", "Workflow resume with approve/reject", "Governance checkpoints", "Good", "MS enterprise governance"],
        ["Literature miner", "Real pipeline + PubMed", "Core use case", "Real", "GA"],
        ["Hypothesis generation", "Dedicated service + LLM", "Discovery Engine core", "Real", "GA"],
        ["Target discovery", "target_discovery_service", "Drug target identification agents", "Real", "GA"],
        ["RAG / Q&A", "rag_service + ChromaDB", "NL over knowledge bases", "Real", "GA"],
        ["Experiment planner", "experiment_service", "Experiment design agents", "Real", "GA"],
        ["Scientific reports", "Two-pass LLM + PDF/DOCX export", "Analysis / reporting in loop", "Real", "GA"],
        ["KG Builder agent", "knowledge_agent_service", "Bookshelf + graph agents", "Partial", ""],
        ["Ontology mapper", "knowledge_agent_service", "KB curation", "Partial", ""],
        ["Cheminformatics (RDKit)", "ADMET, docking hooks", "Chemistry tools + partner FMs", "Real", "GA + partners"],
        ["SLM routing", "Mistral Ministral 8B for light tasks", "Foundry model deployments", "Real", ""],
        ["Custom tools", "Tool Fabric registry + HTTP tools", "Extensible tool framework on Supercomputer", "Good", "MS HPC-scale tools"],
        ["Third-party foundation models", "BYO via API keys", "Partner models (e.g. Insilico) on platform", "BYO", "GA ecosystem"],
        ["Lab / robotics integration", "Not implemented", "Future: lab procedures, IoT, robotics", "None", "MS roadmap"],
        ["Mock / placeholder agents", "~50 agents return mock output", "N/A — production platform", "Gap", "SciNova demo risk if wrong agent picked"],
    ]
    write_table(ws, headers, rows)
    set_col_widths(ws, [28, 36, 36, 14, 30])


def sheet_compute_governance(wb: Workbook) -> None:
    ws = wb.create_sheet("Compute & Governance")
    headers = ["Area", "SciAi-Nova OS", "Microsoft Discovery", "Winner for enterprise", "Winner for TCS demo"]
    rows = [
        ["HPC cluster", "None", "Supercomputer on AKS", "Microsoft", "N/A"],
        ["Molecular simulation at scale", "Not supported", "Supported via tools", "Microsoft", "N/A"],
        ["Single-VM deploy", "Yes — DS3_v2 sufficient", "Not applicable (distributed Azure)", "N/A", "SciNova"],
        ["Elastic compute", "Manual VM resize", "AKS auto-scale", "Microsoft", "N/A"],
        ["Authentication", "JWT + bcrypt users", "Azure AD / enterprise IAM", "Microsoft", "Tie"],
        ["RBAC", "Roles: admin, scientist, reviewer + projects", "Enterprise RBAC, workspace policies", "Microsoft", "Tie"],
        ["Audit trail", "AuditEvent table + UI", "Enterprise auditability", "Microsoft", "SciNova sufficient for demo"],
        ["GxP checks", "governance UI + gxp-check endpoint", "Compliance frameworks", "Microsoft", "SciNova shows concept"],
        ["Risk alerts & approvals", "Built-in UI", "Governance controls", "Tie", "Tie"],
        ["Data residency control", "You choose Azure region / AWS region", "Azure regions", "Tie", "SciNova (multi-cloud)"],
        ["Enterprise SLA", "None (self-operated)", "Microsoft enterprise support", "Microsoft", "N/A"],
        ["Preview / GA status", "Internal / pilot GA (your deployments)", "Cloud GA; app preview", "Microsoft", "SciNova (available now)"],
        ["Security certifications", "Depends on host VM hardening", "Azure + platform compliance story", "Microsoft", "N/A"],
    ]
    write_table(ws, headers, rows)
    set_col_widths(ws, [24, 34, 34, 18, 18])


def sheet_cost(wb: Workbook) -> None:
    ws = wb.create_sheet("Cost & Licensing")
    headers = ["Item", "SciAi-Nova OS", "Microsoft Discovery Cloud", "Microsoft Discovery App"]
    rows = [
        ["Platform license fee", "None (open deployment)", "Consumption-based + enterprise agreement possible", "Free download"],
        ["Infrastructure", "Azure DS3_v2 ~$70–100/mo OR AWS EC2 similar", "AKS, storage, networking — variable, typically $$$", "Local machine only"],
        ["LLM costs", "OpenAI + Mistral API usage (you control)", "Azure OpenAI / Foundry tokens", "GitHub Copilot credits"],
        ["Discovery message billing", "N/A", "Per billable user message (API actions)", "N/A"],
        ["Neo4j / Postgres / Redis", "Included in Docker (VM cost)", "Billed as Azure resources", "N/A"],
        ["Hidden costs", "Engineer time for ops, git pull, .env", "IT, Azure admin, Discovery Studio setup", "Copilot subscription"],
        ["Stop when not demoing", "Stop/deallocate VM — pay disk only", "Deallocate Azure resources", "No cloud cost"],
        ["Typical pilot budget", "Low hundreds $/month", "Thousands+ $/month (estimate)", "Copilot tier dependent"],
        ["Free tier available?", "No — but no license fee", "No for cloud platform", "Yes (app download + Copilot)"],
    ]
    write_table(ws, headers, rows)
    set_col_widths(ws, [28, 38, 38, 32])


def sheet_pharma_use_cases(wb: Workbook) -> None:
    ws = wb.create_sheet("Pharma Use Cases")
    headers = [
        "Use case", "SciAi-Nova OS", "Microsoft Discovery",
        "Recommended platform", "Demo script (SciNova)",
    ]
    rows = [
        [
            "Upload disease review papers (MASH)",
            "Data Fabric → indexed in minutes",
            "Bookshelf KB creation",
            "SciNova for speed",
            "MASH-417 project → 2 PDFs → indexed",
        ],
        [
            "Semantic search over corpus",
            "Fabric search + RAG",
            "NL over Knowledge Base",
            "Both",
            "Search GLP-1 MASH fibrosis",
        ],
        [
            "Knowledge graph exploration",
            "KG Explorer 3D; SQL graph for project data",
            "Investigations + graph reasoning",
            "SciNova for demo UI",
            "SQL graph → MASLD search",
        ],
        [
            "Target identification workflow",
            "Target Discovery pipeline (real)",
            "Biology/pharma agents",
            "Both; SciNova ready today",
            "Run Target Discovery on MASH query",
        ],
        [
            "Hypothesis generation",
            "Hypothesis service + workflow",
            "Discovery Engine core",
            "Both",
            "Literature → Hypothesis workflow",
        ],
        [
            "Experiment design",
            "Experiment planner agent",
            "Experiment design agents",
            "Both",
            "Experiment planner with MASH context",
        ],
        [
            "Scientific report generation",
            "Reports module + PDF/DOCX",
            "Analysis in discovery loop",
            "SciNova polished for export",
            "Generate study report from workflow",
        ],
        [
            "Molecular property / ADMET",
            "RDKit in-platform",
            "Chemistry tools + partner FMs",
            "Microsoft at scale; SciNova for quick ADMET",
            "Cheminformatics agent",
        ],
        [
            "Large-scale docking / MD",
            "Not supported",
            "HPC Supercomputer",
            "Microsoft only",
            "N/A on SciNova",
        ],
        [
            "Enterprise LIMS + ELN federation",
            "LIMS plate sync (limited); upload exports",
            "Fabric / virtual data integration",
            "Microsoft long-term",
            "Upload CSV; mention roadmap",
        ],
        [
            "Regulatory / GxP audit story",
            "Audit UI + approvals",
            "Enterprise compliance",
            "Microsoft production; SciNova narrative",
            "Show governance dashboard",
        ],
        [
            "Partner demo on AWS (Bedrock only)",
            "Bedrock Claude + Titan — no OpenAI",
            "Azure-centric",
            "SciNova",
            "AWS EC2 preflight-aws.sh",
        ],
    ]
    write_table(ws, headers, rows)
    set_col_widths(ws, [28, 32, 32, 22, 36])


def sheet_feature_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("Feature Matrix")
    headers = [
        "#", "Category", "Feature", "SciAi-Nova OS",
        "Microsoft Discovery", "SciNova (Y/N/Partial)", "Microsoft (Y/N/Partial)",
    ]
    features = [
        ("Platform", "Self-hosted deployment", "Docker Compose one-command stack", "Azure ARM + Studio", "Y", "Y"),
        ("Platform", "Desktop application", "Browser UI only", "Discovery app (Windows)", "Partial", "Y"),
        ("Platform", "Multi-cloud (AWS + Azure)", "Same repo on EC2 and Azure VM", "Azure only", "Y", "N"),
        ("Platform", "Full source code customization", "GitHub repo", "API/extension model", "Y", "Partial"),
        ("Data", "PDF ingestion", "Celery pipeline", "Bookshelf", "Y", "Y"),
        ("Data", "Vector semantic index", "ChromaDB", "GraphRAG vector layer", "Y", "Y"),
        ("Data", "Property graph", "Neo4j + SQL", "KG in Bookshelf/memory", "Y", "Y"),
        ("Data", "GraphRAG unified index", "No", "Yes", "N", "Y"),
        ("Data", "Project / workspace isolation", "Projects + X-Project-Id", "Discovery projects", "Y", "Y"),
        ("AI", "Frontier LLM", "GPT-4o-mini / Bedrock Claude", "Azure Foundry / OpenAI", "Y", "Y"),
        ("AI", "Small language model", "Mistral Ministral 8B", "Via Foundry deployments", "Y", "Y"),
        ("AI", "Agent marketplace UI", "80+ agents listed", "Agent catalog in platform", "Y", "Y"),
        ("AI", "Real agent pipelines (>20)", "~24 working", "Production agents", "Partial", "Y"),
        ("AI", "Discovery Engine / hypothesis tree", "Workflow orchestrator", "Discovery Engine", "Partial", "Y"),
        ("AI", "Multi-agent live UI", "WorkflowBuilder animations", "Discovery Studio", "Y", "Y"),
        ("AI", "Human-in-the-loop approval", "Workflow resume", "Governance checkpoints", "Y", "Y"),
        ("AI", "RAG with citations", "rag_service", "Grounded NL answers", "Y", "Y"),
        ("Integrations", "PubMed", "Yes", "Literature in loop", "Y", "Y"),
        ("Integrations", "KEGG", "Yes", "Via tools", "Y", "Y"),
        ("Integrations", "RDKit cheminformatics", "Yes", "Via chemistry tools", "Y", "Y"),
        ("Integrations", "Microsoft 365", "No", "Yes", "N", "Y"),
        ("Integrations", "Microsoft Fabric", "No", "Yes", "N", "Y"),
        ("Integrations", "Custom HTTP tools", "Tool Fabric", "Platform tools framework", "Y", "Y"),
        ("Compute", "HPC / batch simulation", "No", "Supercomputer AKS", "N", "Y"),
        ("Compute", "Partner foundation models", "BYO API", "Integrated (e.g. Insilico)", "Partial", "Y"),
        ("KG UI", "3D force-directed graph", "ForceGraph3D", "Studio visualizations", "Y", "Partial"),
        ("KG UI", "Graph source toggle", "SQL / Neo4j / Auto", "KB-backed", "Y", "Y"),
        ("Governance", "Audit log", "Yes", "Yes", "Y", "Y"),
        ("Governance", "GxP check UI", "Yes", "Enterprise compliance", "Partial", "Y"),
        ("Governance", "Enterprise SLA", "No", "Yes", "N", "Y"),
        ("Ops", "Preflight scripts", "preflight-aws/azure.sh", "Azure portal setup", "Y", "Partial"),
        ("Ops", "One-VM demo cost", "~$70–100/mo", "High", "Y", "N"),
        ("Pharma", "Drug discovery command center", "Dashboard home", "Investigations home", "Y", "Y"),
        ("Pharma", "Research value chain navigation", "Research Value Chain page", "Domain workflows", "Y", "Y"),
        ("Pharma", "Meeting briefs / collaboration", "Collaboration module", "M365 integration path", "Y", "Partial"),
    ]
    rows = []
    for i, (cat, feat, sci, ms, sy, my) in enumerate(features, 1):
        rows.append([i, cat, feat, sci, ms, sy, my])
    write_table(ws, headers, rows)
    set_col_widths(ws, [5, 14, 28, 32, 32, 12, 12])


def sheet_personas(wb: Workbook) -> None:
    ws = wb.create_sheet("Personas")
    headers = ["Persona", "Goals", "SciAi-Nova OS value", "Microsoft Discovery value", "Talking point"]
    rows = [
        [
            "Scientist / researcher",
            "Faster hypotheses, literature synthesis, trusted citations",
            "Upload PDFs, RAG, KG search, Target Discovery in one UI",
            "NL reasoning partner, Discovery Engine, explainable outcomes",
            "SciNova: evidence in hours from your papers",
        ],
        [
            "R&D IT / platform engineer",
            "Secure deploy, observability, integration",
            "Docker, .env, preflight scripts, full repo",
            "Azure ARM, Discovery Studio, enterprise IAM",
            "SciNova: same stack on Azure or AWS",
        ],
        [
            "Innovation / business sponsor",
            "Time to insight, ROI, governance",
            "Low-cost pilot, measurable demo workflows",
            "Repeatable discovery capability at scale",
            "SciNova proves workflow; MS scales it",
        ],
        [
            "Compliance / quality",
            "Audit trail, GxP alignment",
            "Audit events, approvals, risk alerts UI",
            "Certifications, centralized governance",
            "SciNova shows controls; MS certifies at scale",
        ],
        [
            "TCS demo lead",
            "Partner-ready MASH story next week",
            "MASH-417 project, Bedrock AWS path, branded UI",
            "Requires Azure enterprise setup",
            "SciNova is the demo vehicle",
        ],
    ]
    write_table(ws, headers, rows)
    set_col_widths(ws, [22, 32, 36, 36, 32])


def sheet_references(wb: Workbook) -> None:
    ws = wb.create_sheet("References")
    ws["A1"] = "Sources & disclaimers"
    ws["A1"].font = Font(bold=True, size=12)
    rows = [
        ["Microsoft Discovery overview", "https://learn.microsoft.com/en-us/azure/microsoft-discovery/overview-what-is-microsoft-discovery"],
        ["Discovery vs Discovery app", "https://learn.microsoft.com/en-us/azure/microsoft-discovery/concept-discovery-and-discovery-app"],
        ["Discovery billing", "https://learn.microsoft.com/en-us/azure/microsoft-discovery/concept-discovery-billing"],
        ["SciAi-Nova OS repo", "https://github.com/saurabhdsh/scinova-os"],
        ["SciAi-Nova design doc", "design.md in repository"],
        ["Agent status", "AGENT_SPEC_CARDS.md — ~24/78 agents with real pipelines"],
        ["Disclaimer", "SciNova maturity based on codebase as of 2026; Microsoft features per public docs. Compare for positioning, not procurement."],
    ]
    for i, (k, v) in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    set_col_widths(ws, [28, 80])


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet_executive_summary(wb)
    sheet_architecture(wb)
    sheet_knowledge_data(wb)
    sheet_agents(wb)
    sheet_compute_governance(wb)
    sheet_cost(wb)
    sheet_pharma_use_cases(wb)
    sheet_feature_matrix(wb)
    sheet_personas(wb)
    sheet_references(wb)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
